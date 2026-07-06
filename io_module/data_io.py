# -*- coding: utf-8 -*-
"""
Модуль загрузки и сохранения данных.
Абстрагирует работу с различными форматами файлов.
"""

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from loguru import logger
from typing import List, Optional, Tuple
from config.settings import (OUTPUT_DATA_DIR,
                             OSV_GENERAL_DIR,
                             ACCOUNTS_OSV_DIR,
                             REFERENCE_DATA_FILE,
                             ACCOUNTS_OSV_LEASE_DIR,
                             ACCOUNT_CARDS_DIR)
from data_processors import FileHandler


class DataLoader:
    """
    Класс для загрузки данных из различных источников.
    
    Принципы:
    - Все методы следуют единому стилю
    - Общая логика вынесена в приватные методы
    - Валидация выполняется централизованно
    """
    
    # =========================================================================
    # ПРИВАТНЫЕ ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ
    # =========================================================================
    
    @staticmethod
    def _validate_file(path: Path) -> None:
        """
        Проверяет существование файла и его формат.
        
        Raises:
            FileNotFoundError: Если файл не существует
            ValueError: Если файл не .xlsx
        """
        if not path.is_file():
            raise FileNotFoundError(f"Файл {path.name} не существует")
        if path.suffix.lower() != '.xlsx':
            raise ValueError(f"Файл {path.name} не является .xlsx")
    
    @staticmethod
    def _validate_directory(path: Path, pattern: str = "*.xlsx") -> List[Path]:
        """
        Проверяет существование директории и наличие файлов по паттерну.
        
        Args:
            path: Путь к директории
            pattern: Паттерн для поиска файлов (glob-синтаксис)
            
        Returns:
            Список найденных файлов
            
        Raises:
            FileNotFoundError: Если директория не существует или в ней нет файлов
            NotADirectoryError: Если путь не является директорией
        """
        # 1. Проверка существования
        if not path.exists():
            raise FileNotFoundError(f"Директория {path} не существует")
        
        # 2. Проверка, что это директория
        if not path.is_dir():
            raise NotADirectoryError(f"{path} не является директорией")
        
        # 3. Поиск файлов по паттерну
        files = sorted(path.glob(pattern))
        
        # 4. Проверка наличия файлов
        if not files:
            raise FileNotFoundError(
                f"В директории {path} не найдено файлов по паттерну '{pattern}'. "
                f"Проверьте, что выгрузки из 1С размещены в правильной папке."
            )
        
        return files
    
    @staticmethod
    def _load_raw_excel(path: Path, is_header=True) -> pd.DataFrame:
        """
        Загружает сырой Excel-файл с исправлением регистра 1С.
        
        Это базовая операция I/O — только чтение, без бизнес-логики.
        
        Args:
            path: Путь к .xlsx файлу
            is_header: устанавливать ли шапку
            
        Returns:
            Сырой DataFrame
            
        Raises:
            FileNotFoundError: Если файл не существует
            ValueError: Если файл не .xlsx или не читается
        """
        DataLoader._validate_file(path)
        
        try:
            handler = FileHandler()
            fixed_stream = handler._fix_1c_excel_case(path)  # TODO: сделать публичным
            if is_header:
                return pd.read_excel(fixed_stream)
            else:
                return pd.read_excel(fixed_stream, header=None)
        except Exception as e:
            raise ValueError(f"Ошибка чтения файла {path.name}: {e}") from e
    
    @staticmethod
    def _process_with_handler(
        input_path: Path, 
        type_register: str
    ) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Обрабатывает файл(ы) через FileHandler и возвращает результат.
        
        Args:
            input_path: Путь к файлу или папке
            type_register: Тип регистра ('accountosv', 'generalosv', 'analisys', 'posting')
            
        Returns:
            Кортеж (df, check_df)
        """
        handler = FileHandler()
        result = handler.handle_input(
            input_path=input_path,
            type_register=type_register
        )
        
        if type_register not in result:
            raise ValueError(f"FileHandler не вернул результат для '{type_register}'")
        
        df, check_df = result[type_register]
        
        # Логируем проблемные файлы
        if handler.not_correct_files:
            logger.warning(
                f"Проблемы при обработке {len(handler.not_correct_files)} файл(ов):\n" +
                "\n".join(f"  - {name}: {error}" for name, error in handler.not_correct_files.items())
            )
        
        return df, check_df
    
    @staticmethod
    def _find_single_file(
        folder: Path, 
        pattern: str,
        description: str
    ) -> Path:
        """
        Ищет единственный файл по паттерну в папке.
        
        Args:
            folder: Папка для поиска
            pattern: Глоб-паттерн (например, "*_общаяосв_*.xlsx")
            description: Описание для логирования и ошибок
            
        Returns:
            Путь к найденному файлу
            
        Raises:
            FileNotFoundError: Если папка не существует или файлов нет
        """
        if not folder.exists() or not folder.is_dir():
            raise FileNotFoundError(f"Папка {folder} не существует или не является директорией")
        
        files = sorted(folder.glob(pattern))
        if not files:
            raise FileNotFoundError(
                f"В папке {folder} не найдено файлов по паттерну '{pattern}' ({description})"
            )
        
        return files[0]
    
    # =========================================================================
    # ПУБЛИЧНЫЕ МЕТОДЫ ЗАГРУЗКИ — ЕДИНЫЙ СТИЛЬ
    # =========================================================================
    # Все методы принимают путь (Path) и возвращают DataFrame или кортеж.
    # Методы без параметра path — это высокоуровневые обертки, которые
    # сначала находят файл, а потом вызывают базовый метод.
    # =========================================================================
    
    # ----- Спецотчеты -----
    @staticmethod
    def load_process_60_invest(path: Path) -> pd.DataFrame:
        """
        Загружает спецотчет 'ОСВ по 60 ИнвестДоговоры'.
        
        Args:
            path: Путь к файлу ОСВ 76.07
            
        Returns:
            Кортеж (df, check_df) — данные и контрольная таблица
        """
        logger.info(f"Загрузка ОСВ по 60 ИнвестДоговоры: {path.name}")
        
        DataLoader._validate_file(path)
        df, check_df = DataLoader._process_with_handler(path, 'accountosv')
        
        if df.empty:
            raise ValueError(f"ОСВ по 60 ИнвестДоговоры {path.name} не содержит данных")
        
        logger.info(f"Загружен ОСВ по 60 ИнвестДоговоры: {len(df)} строк")
        return df, check_df
    
    
    @staticmethod
    def process_depreciation_statement_decoding(path: Path) -> pd.DataFrame:
        """
        Загружает спецотчет 'Ведомость амортизации'.
        
        Args:
            path: Путь к файлу спецотчета
            
        Returns:
            Сырой DataFrame спецотчета (без бизнес-обработки)
        """
        logger.info(f"Загрузка Ведомость амортизации: {path.name}")
        
        df = DataLoader._load_raw_excel(path)
        
        if df.empty:
            raise ValueError(f"Спецотчет {path.name} не содержит данных")
        
        logger.info(f"Загружена Ведомость амортизации: {len(df)} строк")
        return df
    
    @staticmethod
    def load_lease_report_decoding(path: Path) -> pd.DataFrame:
        """
        Загружает спецотчет 'Расшифровка строк 1450, 1550, 1230 
        по арендным/лизинговым договорам'.
        
        Args:
            path: Путь к файлу спецотчета
            
        Returns:
            Сырой DataFrame спецотчета (без бизнес-обработки)
        """
        logger.info(f"Загрузка расшифровки строк 1450/1550/1230: {path.name}")
        
        df = DataLoader._load_raw_excel(path)
        
        if df.empty:
            raise ValueError(f"Спецотчет {path.name} не содержит данных")
        
        logger.info(f"Загружена расшифровка 1450/1550/1230: {len(df)} строк")
        return df
    
    @staticmethod
    def load_long_short_register(path: Path) -> pd.DataFrame:
        """
        Загружает спецотчет 'Долгие/короткие активы по 97 счету'.
        
        Args:
            path: Путь к файлу спецотчета
            
        Returns:
            Сырой DataFrame спецотчета (без бизнес-обработки)
        """
        logger.info(f"Загрузка спецотчета долг/кор по 97: {path.name}")
        
        df = DataLoader._load_raw_excel(path)
        
        if df.empty:
            raise ValueError(f"Спецотчет {path.name} не содержит данных")
        
        logger.info(f"Загружен спецотчет долг/кор по 97: {len(df)} строк")
        return df
    
    # ----- АНАЛИЗ СЧЕТА -----
    
    @staticmethod
    def load_84_analysis(path: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Загружает и обрабатывает анализ по счету 84.
        
        Args:
            path: Путь к файлу Анализ 84
            
        Returns:
            Кортеж (df, check_df) — данные и контрольная таблица
        """
        logger.info(f"Загрузка анализ 84: {path.name}")
        
        DataLoader._validate_file(path)
        df, check_df = DataLoader._process_with_handler(path, 'analisys')
        
        if df.empty:
            raise ValueError(f"Анализ 84 {path.name} не содержит данных")
        
        logger.info(f"Загружен Анализ 84: {len(df)} строк")
        return df, check_df
    
    # ----- ОСВ -----
    
    @staticmethod
    def load_account_osv76_lease() -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Загружает и обрабатывает ОСВ по субсчетам 76 в части аренды/лизинга (договоры).
        
        Загружает ВСЕ .xlsx файлы из папки ACCOUNTS_OSV_LEASE_DIR,
        объединяет их в один DataFrame.
        
        Returns:
            Кортеж (df, check_df) — данные и контрольная таблица
        """
        logger.info(f"Загрузка ОСВ 76 Аренда: {ACCOUNTS_OSV_LEASE_DIR.name}")
        
        # ★ ВАЛИДАЦИЯ ДИРЕКТОРИИ (а не файла)
        files = DataLoader._validate_directory(ACCOUNTS_OSV_LEASE_DIR, "*.xlsx")
        logger.debug(f"Найдено {len(files)} файл(ов) для загрузки: {[f.name for f in files]}")
        
        # Загрузка и объединение всех файлов
        df, check_df = DataLoader._process_with_handler(ACCOUNTS_OSV_LEASE_DIR, 'accountosv')
        
        if df.empty:
            raise ValueError(
                f"ОСВ 76 Аренда ({ACCOUNTS_OSV_LEASE_DIR.name}) "
                f"не содержит данных после обработки {len(files)} файл(ов)"
            )
        
        logger.info(f"Загружена ОСВ 76 Аренда: {len(df)} строк из {len(files)} файл(ов)")
        return df, check_df
    
    @staticmethod
    def load_transaction_report() -> pd.DataFrame:
        """
        Загружает и объединяет все Отчеты по проводкам из INPUT_DATA.
        
        Это высокоуровневый метод — сам находит файлы в папке.
        
        Returns:
            Сводный DataFrame по всем отчетам по проводкам
        """
        logger.info("Загрузка сводного отчета по проводкам по счетам")
        
        # Поиск файлов
        if not ACCOUNT_CARDS_DIR.exists() or not ACCOUNT_CARDS_DIR.is_dir():
            raise FileNotFoundError(f"Папка {ACCOUNT_CARDS_DIR} не существует")
        
        transaction_report_files = sorted(ACCOUNT_CARDS_DIR.glob("*_отчпровод_*.txt"))
        if not transaction_report_files:
            raise FileNotFoundError(f"В папке {ACCOUNT_CARDS_DIR} нет файлов с отчетами по проводкам (*_отчпровод_*.txt)")
        
        logger.info(f"Найдено {len(transaction_report_files)} файлов отчеты по проводкам")
        
        # Обработка
        df, _ = DataLoader._process_with_handler(ACCOUNT_CARDS_DIR, 'posting')
        
        if df.empty:
            raise ValueError("Сводный Отчет по проводкам не содержит данных")
        
        logger.info(f"Загружен сводный Отчет по проводкам: {len(df)} строк, {len(df.columns)} столбцов")
        return df
    
    @staticmethod
    def load_account_osv() -> pd.DataFrame:
        """
        Загружает и объединяет все ОСВ по счетам из INPUT_DATA.
        
        Это высокоуровневый метод — сам находит файлы в папке.
        
        Returns:
            Сводный DataFrame по всем ОСВ
        """
        logger.info("Загрузка сводной ОСВ по счетам")
        
        # Поиск файлов
        if not ACCOUNTS_OSV_DIR.exists() or not ACCOUNTS_OSV_DIR.is_dir():
            raise FileNotFoundError(f"Папка {ACCOUNTS_OSV_DIR} не существует")
        
        osv_files = sorted(ACCOUNTS_OSV_DIR.glob("*_осв_*.xlsx"))
        if not osv_files:
            raise FileNotFoundError(f"В папке {ACCOUNTS_OSV_DIR} нет файлов ОСВ (*_осв_*.xlsx)")
        
        logger.info(f"Найдено {len(osv_files)} файлов ОСВ")
        
        # Обработка
        df, _ = DataLoader._process_with_handler(ACCOUNTS_OSV_DIR, 'accountosv')
        
        if df.empty:
            raise ValueError("Сводная ОСВ не содержит данных")
        
        logger.info(f"Загружена сводная ОСВ: {len(df)} строк, {len(df.columns)} столбцов")
        return df
    
    @staticmethod
    def load_general_osv() -> Tuple[pd.DataFrame, str]:
        """
        Загружает общую ОСВ из INPUT_DATA.
        
        Это высокоуровневый метод — сам находит файл в папке.
        
        Returns:
            Кортеж (df, filename) — данные и имя исходного файла
        """
        logger.debug("Загрузка общей ОСВ")
        
        # Поиск файла
        file_path = DataLoader._find_single_file(
            OSV_GENERAL_DIR, 
            "*_общаяосв_*.xlsx",
            "общая ОСВ"
        )
        
        # Обработка
        df, _ = DataLoader._process_with_handler(file_path, 'generalosv')
        
        if df.empty:
            raise ValueError(f"Общая ОСВ {file_path.name} не содержит данных")
        
        logger.info(f"Загружена общая ОСВ: {file_path.name} ({len(df)} строк)")
        return df, file_path.name
    
    # ----- Справочники -----
    
    @staticmethod
    def load_reference_data(
        sheet_name: str,
        strings: Optional[List[str]] = None,
        usecols: Optional[List[int]] = None
    ) -> pd.DataFrame:
        """
        Загружает справочные данные из файла справочников.
        
        Args:
            sheet_name: Имя листа в файле справочников
            strings: Столбцы, которые нужно привести к типу 'string'
            usecols: Индексы столбцов для чтения (None = все)
            
        Returns:
            DataFrame со справочными данными
        """
        logger.debug(f"Загрузка справочника: {sheet_name}")
        
        if not REFERENCE_DATA_FILE.exists():
            raise FileNotFoundError(f"Файл справочников {REFERENCE_DATA_FILE} не найден")
        
        try:
            df = pd.read_excel(
                REFERENCE_DATA_FILE,
                sheet_name=sheet_name,
                dtype='string',
                usecols=usecols
            )
            
            if strings:
                existing_cols = [col for col in strings if col in df.columns]
                if existing_cols:
                    df = df.astype({col: 'string' for col in existing_cols})
            
            logger.debug(f"Загружен справочник '{sheet_name}': {len(df)} строк")
            return df
            
        except Exception as e:
            raise ValueError(f"Ошибка загрузки справочника '{sheet_name}': {e}") from e

class DataSaver:
    """Класс для сохранения результатов обработки."""
    
    # =========================================================================
    # ФОРМАТИРОВАНИЕ EXCEL
    # =========================================================================
    
    @staticmethod
    def _apply_excel_formatting(worksheet, numeric_columns: list = None) -> None:
        """
        Применяет форматирование к листу Excel:
        - Заголовки столбцов — жирным шрифтом с выравниванием по центру
        - Числовые столбцы — 2 знака после запятой с разделителями разрядов
        - Автоматическая ширина столбцов
        
        Args:
            worksheet: Объект листа openpyxl
            numeric_columns: Список имён числовых столбцов (опционально)
        """
        # Стили
        bold_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        number_format = '#,##0.00'
        
        # Определяем числовые столбцы, если не указаны явно
        if numeric_columns is None:
            numeric_columns = []
            for col_idx, cell in enumerate(worksheet[1], start=1):
                # Проверяем вторую строку (первая — заголовок)
                if worksheet.max_row >= 2:
                    value = worksheet.cell(row=2, column=col_idx).value
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        numeric_columns.append(col_idx)
        
        # Форматирование заголовков
        for cell in worksheet[1]:
            cell.font = bold_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        # Форматирование числовых столбцов
        for col_idx in numeric_columns:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format
                    cell.alignment = Alignment(horizontal='right')
        
        # Автоматическая ширина столбцов
        for col_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row in range(1, min(worksheet.max_row + 1, 100)):  # Ограничим для производительности
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            # Ширина = max(длина контента, 10), но не более 50
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Закрепляем первую строку (заголовки)
        worksheet.freeze_panes = 'A2'
    
    # =========================================================================
    # СОХРАНЕНИЕ КОМБИНИРОВАННОГО ОТЧЁТА
    # =========================================================================
    
    @staticmethod
    def save_combined_report(
        final_report: pd.DataFrame,
        main_df: pd.DataFrame,
        filename: str,
        subfolder: str = None
    ) -> Path:
        """
        Сохраняет два DataFrame в один Excel-файл на разных листах.
        
        Args:
            final_report: DataFrame для листа "Расшифровка_ББЛ"
            main_df: DataFrame для листа "исходники"
            filename: Имя выходного файла
            subfolder: Подпапка в OUTPUT_DATA (опционально)
            
        Returns:
            Путь к сохранённому файлу
        """
        output_dir = OUTPUT_DATA_DIR
        if subfolder:
            output_dir = output_dir / subfolder
            output_dir.mkdir(parents=True, exist_ok=True)
        
        output_path = output_dir / filename
        
        # =========================================================================
        # ПОДГОТОВКА final_report: сброс индекса и перемещение столбца
        # =========================================================================
        
        # Сбрасываем индекс, чтобы он стал обычным столбцом
        final_report_reset = final_report.reset_index()
        
        # Перемещаем столбец "Итоговый номер счета" на второе место
        # (между "РСБУ Код отчетности" и "1 уровень")
        if 'Итоговый номер счета' in final_report_reset.columns:
            cols = final_report_reset.columns.tolist()
            cols.remove('Итоговый номер счета')
            
            # Находим позицию для вставки (после "РСБУ Код отчетности")
            if 'РСБУ Код отчетности' in cols:
                insert_idx = cols.index('РСБУ Код отчетности') + 1
            else:
                insert_idx = 1  # Если "РСБУ Код отчетности" нет, вставляем на второе место
            
            cols.insert(insert_idx, 'Итоговый номер счета')
            final_report_reset = final_report_reset[cols]
        
        # =========================================================================
        # ОПРЕДЕЛЕНИЕ ЧИСЛОВЫХ СТОЛБЦОВ ДЛЯ ФОРМАТИРОВАНИЯ
        # =========================================================================
        
        # Для final_report: исключаем индексный столбец из числового форматирования
        final_numeric_cols = [
            i + 1 for i, col in enumerate(final_report_reset.columns)
            if pd.api.types.is_numeric_dtype(final_report_reset[col])
            and col != 'Итоговый номер счета'  # ← Исключаем индекс
        ]
        
        # Для main_df: все числовые столбцы
        main_numeric_cols = [
            i + 1 for i, col in enumerate(main_df.columns)
            if pd.api.types.is_numeric_dtype(main_df[col])
        ]
        
        # =========================================================================
        # СОХРАНЕНИЕ В EXCEL
        # =========================================================================
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Лист 1: Расшифровка ББЛ
                final_report_reset.to_excel(
                    writer, 
                    sheet_name='Расшифровка_ББЛ', 
                    index=False  # ← Индекс уже сброшен, не дублируем
                )
                DataSaver._apply_excel_formatting(
                    writer.sheets['Расшифровка_ББЛ'], 
                    final_numeric_cols
                )
                
                # Лист 2: Исходники
                main_df.to_excel(
                    writer, 
                    sheet_name='исходники', 
                    index=False
                )
                DataSaver._apply_excel_formatting(
                    writer.sheets['исходники'], 
                    main_numeric_cols
                )
        except PermissionError:
            raise PermissionError(
                f"❌ Не удалось сохранить результат в '{output_path.name}': "
                f"файл открыт в Excel. Закройте файл и перезапустите программу."
            )
        
        logger.info(f"Комбинированный отчёт сохранён в {output_path.name}")
        return output_path
    
    # =========================================================================
    # СТАРЫЕ МЕТОДЫ (для обратной совместимости)
    # =========================================================================
    
    @staticmethod
    def save_to_excel(df: pd.DataFrame, filename: str, subfolder: str = None) -> Path:
        """
        Сохраняет DataFrame в Excel файл с форматированием.
        
        Применяет:
        - Жирные заголовки с выравниванием по центру
        - Числовой формат с разделителями разрядов (2 знака после запятой)
        - Автоматическую ширину столбцов
        - Закрепление первой строки
        
        Args:
            df: DataFrame для сохранения
            filename: Имя выходного файла
            subfolder: Подпапка в OUTPUT_DATA (опционально)
            
        Returns:
            Путь к сохранённому файлу
        """
        output_dir = OUTPUT_DATA_DIR
        if subfolder:
            output_dir = output_dir / subfolder
            output_dir.mkdir(parents=True, exist_ok=True)
        
        output_path = output_dir / filename
        
        # Определяем числовые столбцы для форматирования
        numeric_cols = [
            i + 1 for i, col in enumerate(df.columns)
            if pd.api.types.is_numeric_dtype(df[col])
        ]
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Sheet1', index=False)
                DataSaver._apply_excel_formatting(
                    writer.sheets['Sheet1'],
                    numeric_cols
                )
        except PermissionError:
            raise PermissionError(
                f"❌ Не удалось сохранить результат в '{output_path.name}': "
                f"файл открыт в Excel. Закройте файл и перезапустите программу."
            )
        
        logger.debug(f"Результаты сохранены в {output_path.name}")
        return output_path
