# -*- coding: utf-8 -*-
"""
Created on Tue Jun 16 16:32:25 2026

@author: a.karabedyan
"""

# -*- coding: utf-8 -*-
"""
Created on Tue Apr 21 14:01:44 2026

@author: a.karabedyan
"""
from abc import ABC, abstractmethod
from io import BytesIO
from typing import List

import openpyxl
import xlwings as xw
import pandas as pd
import re

import tempfile
import os

import time

# Компилируем regex один раз для скорости
_ACCOUNT_PATTERN = re.compile(r'^\d{2}(\.\d{1,2})*$')

'''
Строки с данными значениями будут удалены из регистров,
чтобы исключить дублирующие или промежуточные обороты
'''
exclude_values = ['Нач.сальдо',
                  'Оборот',
                  'Итого оборот',
                  'Кон.сальдо',
                  'Начальное сальдо',
                  'Конечное сальдо',
                  'Кор. Субконто1',
                  'Кол-во:',
                  'Итого',
                  'Количество',

                  ]


DESIRED_ORDER = {'card':{
                    'upp': [
                        'Имя_файла', 'Дата', 'Документ', 'Дебет', 'Дебет_значение', 'Кредит', 
                        'Кредит_значение', 'Текущее сальдо', 'Текущее сальдо_значение', 
                        'Операция_1', 'Операция_2', 'Операция_3', 'Операция_4', 'Операция_5', 
                        'Операция_6', 'Дт_количество', 'Кт_количество', 'Дт_валюта', 
                        'Дт_валютая_сумма', 'Кт_валюта', 'Кт_валютая_сумма'
                    ],
                    'not_upp': [
                        'Имя_файла', 'Период', 'Дебет', 'Дебет_значение', 'Кредит', 
                        'Кредит_значение', 'Текущее сальдо', 'Текущее сальдо_значение', 
                        'Документ_1', 'Документ_2', 'Аналитика Дт_1', 'Аналитика Дт_2', 
                        'Аналитика Дт_3', 'Аналитика Дт_4', 'Аналитика Кт_1', 'Аналитика Кт_2', 
                        'Аналитика Кт_3', 'Аналитика Кт_4', 'Дебет_количество', 
                        'Кредит_количество', 'Дебет_валюта', 'Дебет_валютное_количество', 
                        'Кредит_валюта', 'Кредит_валютное_количество'
                    ]},
                'posting':{
                    'upp': [
                        'Имя_файла', 'Дата', 'Документ', 'Дт', 'Кт', 'Сумма', 
                        'Содержание_1', 'Содержание_2', 'Содержание_3',
                        'Субконто Дт1', 'Субконто Дт2', 'Субконто Дт3',
                        'Субконто Кт1', 'Субконто Кт2', 'Субконто Кт3',
                        'Дт_количество', 'Кт_количество',
                        'Дт_валюта', 'Дт_валюта_количество',
                        'Кт_валюта', 'Кт_валюта_количество'
                        ],
                    'not_upp': [
                        'Имя_файла', 'Период', 'Дебет', 'Дебет_значение', 'Кредит', 'Кредит_значение',
                        'Документ_1', 'Документ_2',
                        'Аналитика Дт_1', 'Аналитика Дт_2', 'Аналитика Дт_3', 'Аналитика Дт_4',
                        'Аналитика Кт_1', 'Аналитика Кт_2', 'Аналитика Кт_3', 'Аналитика Кт_4',
                        'Дебет_количество', 'Кредит_количество',
                        'Дебет_валюта', 'Дебет_валютное_количество',
                        'Кредит_валюта', 'Кредит_валютное_количество'
                    ]},
                'analisys':{
                    'upp': [
                        'Исх.файл', 'Субсчет', 'Аналитика', 'Корр_счет', 'Субконто_корр_счета', 'Вид связи КА за период','С кред. счетов', 
                        'С кред. счетов_КОЛ', 'С кред. счетов_ВАЛ', 'В дебет счетов', 'В дебет счетов_КОЛ', 'В дебет счетов_ВАЛ',
                        'Level_0', 'Level_1', 'Level_2'
                        ],
                    'not_upp': [
                        'Исх.файл', 'Субсчет', 'Аналитика', 'Корр_счет', 'Субконто_корр_счета', 'Вид связи КА за период','С кред. счетов', 
                        'С кред. счетов_КОЛ', 'С кред. счетов_ВАЛ', 'В дебет счетов', 'В дебет счетов_КОЛ', 'В дебет счетов_ВАЛ',
                        'Level_0', 'Level_1', 'Level_2'
                    ]},
                'turnover':{
                    'upp': [
                        'Исх.файл', 'Субконто',
                        'Дебет_начало', 'Количество_Дебет_начало', 'ВалютнаяСумма_Дебет_начало',
                        'Кредит_начало', 'Количество_Кредит_начало', 'ВалютнаяСумма_Кредит_начало',
                        'Дебет_оборот', 'Количество_Дебет_оборот', 'ВалютнаяСумма_Дебет_оборот',
                        'Кредит_оборот', 'Количество_Кредит_оборот', 'ВалютнаяСумма_Кредит_оборот',
                        'Дебет_конец', 'Количество_Дебет_конец', 'ВалютнаяСумма_Дебет_конец',
                        'Кредит_конец', 'Количество_Кредит_конец', 'ВалютнаяСумма_Кредит_конец','Валюта',
                        'Начало периода  для вида связи', 'Конец  периода для вида связи', 'Вид связи КА за период',
                        'Level_0', 'Level_1', 'Level_2'
                        ],
                    'not_upp': [
                       'Исх.файл', 'Субконто',
                       'Дебет_начало', 'Количество_Дебет_начало', 'ВалютнаяСумма_Дебет_начало',
                       'Кредит_начало', 'Количество_Кредит_начало', 'ВалютнаяСумма_Кредит_начало',
                       'Дебет_оборот', 'Количество_Дебет_оборот', 'ВалютнаяСумма_Дебет_оборот',
                       'Кредит_оборот', 'Количество_Кредит_оборот', 'ВалютнаяСумма_Кредит_оборот',
                       'Дебет_конец', 'Количество_Дебет_конец', 'ВалютнаяСумма_Дебет_конец',
                       'Кредит_конец', 'Количество_Кредит_конец', 'ВалютнаяСумма_Кредит_конец','Валюта',
                       'Начало периода  для вида связи', 'Конец  периода для вида связи', 'Вид связи КА за период',
                       'Level_0', 'Level_1', 'Level_2'
                    ]},
                'accountosv':{
                    'upp': [
                        'Исх.файл', 'Субконто',
                        'Дебет_начало', 'Количество_Дебет_начало', 'ВалютнаяСумма_Дебет_начало',
                        'Кредит_начало', 'Количество_Кредит_начало', 'ВалютнаяСумма_Кредит_начало',
                        'Дебет_оборот', 'Количество_Дебет_оборот', 'ВалютнаяСумма_Дебет_оборот',
                        'Кредит_оборот', 'Количество_Кредит_оборот', 'ВалютнаяСумма_Кредит_оборот',
                        'Дебет_конец', 'Количество_Дебет_конец', 'ВалютнаяСумма_Дебет_конец',
                        'Кредит_конец', 'Количество_Кредит_конец', 'ВалютнаяСумма_Кредит_конец', 'Валюта',
                        'Начало периода  для вида связи', 'Конец  периода для вида связи', 'Вид связи КА за период',
                        'Level_0', 'Level_1', 'Level_2'
                        ],
                    'not_upp': [
                       'Исх.файл', 'Субконто',
                       'Дебет_начало', 'Количество_Дебет_начало', 'ВалютнаяСумма_Дебет_начало',
                       'Кредит_начало', 'Количество_Кредит_начало', 'ВалютнаяСумма_Кредит_начало',
                       'Дебет_оборот', 'Количество_Дебет_оборот', 'ВалютнаяСумма_Дебет_оборот',
                       'Кредит_оборот', 'Количество_Кредит_оборот', 'ВалютнаяСумма_Кредит_оборот',
                       'Дебет_конец', 'Количество_Дебет_конец', 'ВалютнаяСумма_Дебет_конец',
                       'Кредит_конец', 'Количество_Кредит_конец', 'ВалютнаяСумма_Кредит_конец','Валюта',
                       'Начало периода  для вида связи', 'Конец  периода для вида связи', 'Вид связи КА за период',
                       'Level_0', 'Level_1', 'Level_2'
                    ]},
                'generalosv':{
                    'upp': [
                        'Исх.файл', 'Счет', 'Наименование', 'Дебет_начало', 'Кредит_начало', 'Дебет_оборот', 'Кредит_оборот', 'Дебет_конец', 'Кредит_конец'
                        ],
                    'not_upp': [
                       'Исх.файл', 'Счет', 'Наименование', 'Дебет_начало', 'Кредит_начало', 'Дебет_оборот', 'Кредит_оборот', 'Дебет_конец', 'Кредит_конец'
                       ]}
}

class FileProcessor(ABC):
    """Абстрактный базовый класс для обработчиков файлов"""
    def __init__(self):
        self.table_for_check = pd.DataFrame()  # для хранения данных по оборотам до обработки в таблицах
        self.file = 'NoNameFile'
    
    @staticmethod
    def _is_parent(account: str, accounts: List[str]) -> bool:
        """
        Проверяет, есть ли субсчета у заданного счета.
    
        :param account: Счет, для которого проверяются субсчета.
        :param accounts: Список всех счетов.
        :return: True, если есть хотя бы один субсчет; иначе False.
        """
        return any(acc.startswith(account + '.') and acc != account for acc in accounts)
    
    @staticmethod
    def _is_accounting_code_vectorized(series: pd.Series) -> pd.Series:
        """
        Векторизованная версия для работы с целыми сериями.
        """
        # Конвертируем в строку
        str_series = series.astype(str)
        
        # Быстрые проверки
        result = pd.Series(False, index=series.index)
        
        # Специальные значения
        special_mask = str_series.isin(["0", "00", "000"])
        result[special_mask] = True
        
        # Проверяем наличие точки
        has_dot = str_series.str.contains('.', regex=False)
        
        # Для значений без точки - простые цифровые проверки
        no_dot_mask = ~has_dot
        
        # Проверяем значения без точки
        numeric_no_dot = str_series[no_dot_mask].str.isdigit()
        valid_length_no_dot = str_series[no_dot_mask].str.len() <= 2
        
        # Создаем маски для значений без точки
        valid_numeric_no_dot = pd.Series(False, index=series.index)
        valid_length_mask = pd.Series(False, index=series.index)
        
        # Используем .loc для присвоения значений
        valid_numeric_no_dot.loc[no_dot_mask] = numeric_no_dot
        valid_length_mask.loc[no_dot_mask] = valid_length_no_dot
        
        # Объединяем маски
        valid_no_dot_mask = no_dot_mask & valid_numeric_no_dot & valid_length_mask
        
        # Обновляем результат для значений без точки
        result[valid_no_dot_mask] = True
        
        # Для значений с точкой - сложная проверка
        dot_values = str_series[has_dot]
        if not dot_values.empty:
            # Разделяем на части
            parts = dot_values.str.split('.')
            
            # Проверяем каждую часть
            valid_parts = parts.apply(lambda x: all(
                (p.isdigit() and len(p) <= 2) or (p.isalpha() and len(p) <= 2)
                for p in x if p  # Пропускаем пустые части
            ))
            
            # Проверяем наличие хотя бы одной цифровой части
            has_digit = parts.apply(lambda x: any(p.isdigit() for p in x))
            
            # Приведение к типу bool
            result[has_dot] = (valid_parts & has_digit).to_numpy().astype(bool)
        
        return result

    @staticmethod
    def _preprocessor_openpyxl(file_like_object: BytesIO) -> pd.DataFrame:
        # Ключевые заголовки для поиска проблемной строки
        target_headers = {
            'субконто', 'нач. сальдо деб.', 'нач. сальдо кред.',
            'деб. оборот', 'кред. оборот', 'кон. сальдо деб.', 'кон. сальдо кред.'
        }
    
        # --- Читаем файл через openpyxl ---
        workbook = openpyxl.load_workbook(file_like_object)
        sheet = workbook.active
    
        max_row = sheet.max_row
        max_col = sheet.max_column
    
        # Получаем уровни группировки (для всех строк)
        grouping_levels = [sheet.row_dimensions[row_idx].outline_level for row_idx in range(1, max_row + 1)]
    
        # Ищем столбец с "Кор. Счет" или "Кор.счет" для курсивности
        found_kor_schet_col = None
        for row in sheet.iter_rows(min_row=1, max_row=30, values_only=True):
            for idx, cell_val in enumerate(row, start=1):
                if cell_val in ('Кор. Счет', 'Кор.счет'):
                    found_kor_schet_col = idx
                    break
            if found_kor_schet_col:
                break
    
        # Получаем флаги курсивности по найденному столбцу
        italic_flags = [0] * max_row
        if found_kor_schet_col:
            for row_idx in range(2, max_row + 1):
                cell = sheet.cell(row=row_idx, column=found_kor_schet_col)
                italic_flags[row_idx - 1] = 1 if cell.font and cell.font.italic else 0
    
        # Читаем все данные в список
        data = []
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            data.append(list(row))
    
    
        # Находим индекс строки с заголовками, содержащей все ключевые слова
        header_row_idx = None
        for i in range(min(30, len(data))):
            row_values = [str(cell).strip().lower() if cell is not None else '' for cell in data[i]]
            if target_headers.issubset(set(row_values)):
                header_row_idx = i
                break
    
        # Если нашли проблемную строку — считываем её с xlwings для точного форматирования
        if header_row_idx is not None:
            # --- Считаем проблемную строку с xlwings для точного форматирования ---
            # Сохраняем BytesIO во временный файл (xlwings не работает с BytesIO напрямую)
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                tmp_file.write(file_like_object.getvalue())
                tmp_path = tmp_file.name
    
            try:
                app = xw.App(visible=False)
                wb_xl = app.books.open(tmp_path)
                sht_xl = wb_xl.sheets[0]
    
                # Читаем отформатированные значения проблемной строки (1-based индексация в xlwings)
                formatted_header_row = []
                for col_idx in range(1, max_col + 1):
                    cell = sht_xl.cells(header_row_idx + 1, col_idx)
                    formatted_header_row.append(cell.api.Text)  # именно отображаемое значение
    
                wb_xl.close()
                app.quit()
            finally:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
            # Очистка xlwings для предотвращения "зависших" процессов Excel
            try:
                wb_xl.close()  # Повторно, на случай исключений выше
                app.quit()
            except:
                pass
            # xw.apps.cleanup()
            time.sleep(0.5)  # Паузы для стабилизации
    
            # Заменяем строку в данных на отформатированную из xlwings
            data[header_row_idx] = formatted_header_row
    
        # Формируем DataFrame из всех строк начиная со второй (строка 1 Excel — заголовки или данные, но мы не выделяем)
        # Все строки, включая проблемную, идут в данные DataFrame без выделения заголовка
        df = pd.DataFrame(data[1:], dtype='string', columns=None)  # columns=None, чтобы не использовать какую-либо строку как заголовок
    
        # Добавляем столбцы группировки и курсивности для всех строк DataFrame
        # grouping_levels[1:] — уровни для строк 2 и далее (Excel строки 2+)
        df.insert(0, 'Уровень_группировки', grouping_levels[1:])
        df.insert(1, 'Курсив', italic_flags[1:])
        
        df[['Уровень_группировки', 'Курсив']] = df[['Уровень_группировки', 'Курсив']].apply(pd.to_numeric, errors='coerce')
    
        workbook.close()
        return df

    
    
    @staticmethod
    def _process_dataframe_optimized(df: pd.DataFrame) -> pd.DataFrame:
        """Оптимизированная обработка DataFrame отчёта по проводкам."""
        
        # 1. Поиск строки с "Дата"
        first_col = df.iloc[:, 0].astype(str).str.lower()
        mask = first_col.str.contains('дата', na=False)
        
        if not mask.any():
            raise ValueError('Файл не является регистром 1С (не найдена строка с "Дата")')
        
        date_row_idx = mask.values.argmax()  # ← Безопасно: всегда позиция
        
        # 2. Установка заголовков (с обработкой NaN)
        df.columns = (
            df.iloc[date_row_idx]
            .fillna('')
            .astype(str)
            .str.strip()
        )
        df = df.iloc[date_row_idx + 1:].copy()
        
        # 3. Делаем имена столбцов уникальными
        seen = {}
        unique_columns = []
        for col in df.columns:
            if not col:  # Пустое имя
                col = f'NoNameCol_{len(seen) + 1}'
            if col in seen:
                seen[col] += 1
                unique_columns.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                unique_columns.append(col)
        df.columns = unique_columns
        
        # 4. Проверка обязательных столбцов
        required_cols = ['Дата', 'Документ']
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            raise ValueError(
                f"В выгрузке отсутствуют обязательные столбцы: {missing}. "
                f"Доступные столбцы: {list(df.columns)}"
            )
        
        # 5. Преобразование даты
        df['Дата'] = pd.to_datetime(
            df['Дата'], format='mixed', dayfirst=True, errors='coerce'
        )
        
        # 6. Добавляем суффикс ТОЛЬКО к дубликатам документов
        mask_doc = df['Документ'].notna()
        if mask_doc.any():
            doc_counts = df.loc[mask_doc, 'Документ'].value_counts()
            duplicated_docs = doc_counts[doc_counts > 1].index
            
            dup_mask = mask_doc & df['Документ'].isin(duplicated_docs)
            if dup_mask.any():
                df.loc[dup_mask, 'Документ'] = (
                    df.loc[dup_mask, 'Документ']
                    + '_end'
                    + df.loc[dup_mask].groupby('Документ').cumcount().add(1).astype(str)
                )
        
        # 7. Заполнение пропусков (forward fill для сгруппированных данных)
        df['Дата'] = df['Дата'].ffill()
        df['Документ'] = df['Документ'].ffill()
        
        # 8. Удаление пустых строк и столбцов
        df = df[df['Дата'].notna()].copy()
        df = df.dropna(how='all').dropna(how='all', axis=1)
        
        return df
    
    @classmethod
    def shiftable_level(cls, df: pd.DataFrame) -> pd.DataFrame:
        """
        Выравнивает столбцы таким образом, чтобы бухгалтерские счета находились в одном столбце.
        """
        

        if not df.empty:
            list_lev = [i for i in df.columns.to_list() if 'Level' in i]
            continue_shifting = True
            iteration = 0
            
    
            while continue_shifting:
                continue_shifting = False
                iteration += 1
                previous_table_state = df.copy()
    
                for i in list_lev:
                    # Проверяем, есть ли в столбце и субсчет, и субконто
                    if cls._is_accounting_code_vectorized(df[i]).nunique() == 2:
                        lm = int(i.split('_')[-1])
                        new_list_lev = list_lev[lm:]
                        
                        
                        # Создаем булеву маску для строк, где первый столбец new_list_lev содержит бухгалтерский код
                        
                        mask = cls._is_accounting_code_vectorized(df[new_list_lev[0]])
                        
                        # Применяем сдвиг только к строкам, где mask == True
                        df.loc[mask, new_list_lev] = df.loc[mask, new_list_lev].values
                        
                        # Для остальных строк сдвигаем по-другому (если нужно)
                        if lm > 0:
                            other_cols = list_lev[lm - 1:-1]
                            df.loc[~mask, new_list_lev[:len(other_cols)]] = df.loc[~mask, other_cols].values
                        
                        continue_shifting = True
    
                if previous_table_state.equals(df):
                    break
        return df

    
    def find_max_level_column(self, df: pd.DataFrame) -> str:
        """
        Находит столбец Level_ с максимальным индексом, все значения которого возвращают True
        в методе _is_accounting_code_vectorized.
        """
        # Получаем все столбцы, начинающиеся с 'Level_'
        level_columns = [col for col in df.columns if col.startswith('Level_')]
        
        max_level_column = None
        max_index = -1
    
        for col in level_columns:
            # Проверяем, все ли значения в столбце возвращают True
            if self._is_accounting_code_vectorized(df[col]).all():
                # Извлекаем индекс из имени столбца
                index = int(col.split('_')[1])
                # Обновляем максимальный индекс и соответствующий столбец
                if index > max_index:
                    max_index = index
                    max_level_column = col
    
        return max_level_column


    @abstractmethod
    def process_file(self, file_path: BytesIO) -> pd.DataFrame:
        pass