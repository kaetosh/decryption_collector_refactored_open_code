"""
Created on Mon Jun 22 09:42:06 2026

@author: a.karabedyan
"""

"""
Шаг 1a: Формирование списка необходимых регистров для выгрузки.
"""
import pandas as pd
from loguru import logger

from pipeline.base import Step, ProcessingContext
from pipeline.errors import ReferenceMismatchError
from io_module import DataLoader
from config.settings import REFERENCE_CONFIGS
from utils import process_account, format_filename_vectorized
from config.settings import (
    OUTPUT_DATA_DIR, 
    OPU_ACCOUNTS_PREFIXES,
)


class Step1aListExpectedRegistersStep(Step):
    """
    Шаг 1а: Формирование списка необходимых регистров для выгрузки.
    
    Анализирует общую ОСВ и определяет, какие регистры нужно выгрузить из 1С.
    """
    
    def __init__(self):
        super().__init__(
            name="Шаг 1а: Формирование списка выгрузок",
            description="Определение необходимых регистров на основе общей ОСВ"
        )
    
    @staticmethod
    def _format_worksheet(worksheet) -> None:
        """Базовое форматирование листа Excel."""
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        
        bold_font = Font(bold=True)
        for cell in worksheet[1]:
            cell.font = bold_font
        
        for col_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(worksheet.cell(row=row, column=col_idx).value or ''))
                for row in range(1, worksheet.max_row + 1)
            )
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        worksheet.freeze_panes = 'A2'
    
    def _format_special_reports_sheet(self, worksheet) -> None:
        """
        Специальное форматирование для листа спецотчетов:
        - Добавляет заголовок с пояснением
        - Выделяет цветом строку-предупреждение
        """
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Вставляем заголовок в первую строку
        worksheet.insert_rows(1, 2)
        
        # Строка 1: Заголовок с пояснением
        header_cell = worksheet.cell(
            row=1,
            column=1,
            value=(
                '⚠️ Шаблоны отчетов уточните у админа.'
            )
        )
        header_cell.font = Font(bold=True, size=12, color='C00000')  # тёмно-красный
        header_cell.fill = PatternFill(
            start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'
        )  # светло-жёлтый фон
        header_cell.alignment = Alignment(wrap_text=True)
        
        # Объединяем ячейки заголовка на всю ширину
        worksheet.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=worksheet.max_column
        )
        
        # Строка 2: Пустая (разделитель)
        
        # Форматируем заголовки таблицы (теперь в строке 3)
        bold_font = Font(bold=True)
        for cell in worksheet[3]:
            cell.font = bold_font
        
        # Автоматическая ширина колонок
        for col_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(worksheet.cell(row=row, column=col_idx).value or ''))
                for row in range(3, worksheet.max_row + 1)
            )
            # Для описания делаем шире
            if col_idx == worksheet.max_column:
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 80)
            else:
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Включаем перенос текста для колонки "Что делает отчет"
        for row in range(4, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=worksheet.max_column)
            cell.alignment = Alignment(wrap_text=True)
        
        # Закрепляем заголовок таблицы
        worksheet.freeze_panes = 'A4'
    
    def _build_special_reports_list(
        self,
        name_company: str,
        period: str
    ) -> pd.DataFrame:
        """
        Формирует DataFrame со списком спецотчетов.
        
        Returns:
            DataFrame с колонками: Имя файла, Описание, Обязательность
        """
        reports = [
            {
                'шаблон': f'{name_company}_анализ_84_{period}_.xlsx',
                'описание': (
                    'Анализ 84 счета для годовой отчетности — '
                    'разделяет НРП на результат текущего года и прошлых периодов'
                    'по субсчетам и субсчетам корр счетов'
                ),
                'Обязательность': 'для года - да',
            },
            {
                'шаблон': f'{name_company}_арендареклассдолгкорт_7697_{period}_.xlsx',
                'описание': (
                    'Расшифровка строк 1450, 1550, 1230 по арендным договорам — '
                    'выделяет в арендной задолженности долгую и короткую части'
                ),
                'Обязательность': 'для УПП - да',
            },
            {
                'шаблон': f'{name_company}_лизингреклассдолгкорт_7697_{period}_.xlsx',
                'описание': (
                    'Расшифровка строк 1450, 1550, 1230 по лизинговым договорам — '
                    'выделяет в лизинговой задолженности долгую и короткую части'
                ),
                'Обязательность': 'для УПП - да',
            },
            {
                'шаблон': f'{name_company}_ведамор_0102_{period}_.xlsx',
                'описание': (
                    'Ведомость амортизации ОС — '
                    'разделяет арендованное имущество на "у третьих лиц" и "у компаний группы"'
                ),
                'Обязательность': 'да',
            },
            {
                'шаблон': f'{name_company}_осв_60инвест_{period}_.xlsx',
                'описание': (
                    'ОСВ по 60 счету с признаком "Договор.Инвестиции" — '
                    'выделяет задолженность по внеоборотным активам (ОС)'
                    'Инвестиции (св-во Договоры) -> Контрагенты -> Договоры. Вид взаиморасчетов'
                ),
                'Обязательность': 'для УПП - да',
            },
            {
                'шаблон': f'{name_company}_реклассдолгкорт_97_{period}_.xlsx',
                'описание': (
                    'Расшифровка строк Баланса по долгосрочной и краткосрочной задолженности — '
                    'разделяет РБП на 97.21 на долгосрочную и краткосрочную части'
                ),
                'Обязательность': 'для УПП - да',
            },
        ]
        
        df = pd.DataFrame(reports)
        df = df.rename(columns={'шаблон': 'Имя файла для сохранения'})
        # df['Обязательность'] = 'Необязательно (улучшает детализацию)'
        df['Куда класть'] = '_INPUT_DATA/special_reports/'
        
        # Переупорядочиваем колонки для удобства бухгалтера
        df = df[[
            'Имя файла для сохранения',
            'Куда класть',
            'Обязательность',
            'описание',
        ]]
        df = df.rename(columns={'описание': 'Что делает отчет'})
        
        return df
    
    
    def _process(self, context: ProcessingContext) -> ProcessingContext:
        logger.debug("Проверка наличия всех счетов в справочнике")
        
        # 1. Извлечение данных
        df = context.data['osv'].copy()
        
        # сохраняем ссылку на оригинал для анализа оборотов (ОПУ)
        # (позже df будет очищен от столбцов оборотов)
        self._original_osv_df = df.copy()
        
        name_file = context.get_metadata('osv_filename')
        name_company = context.get_metadata('company_name')
        period = context.get_metadata('period')
        
        # 2. Расчёт сальдо и фильтрация
        df['Сальдо, тыс.ед.'] = (
            df['Дебет_конец']
            .sub(df['Кредит_конец'], fill_value=0)
            .div(1_000)
            .round(2)
        )
        df = df[df['Сальдо, тыс.ед.'].notna() & (df['Сальдо, тыс.ед.'] != 0)]
        
        if df.empty:
            raise ValueError("Нет строк с ненулевым сальдо после фильтрации")
        
        # 3. Очистка от технических колонок
        unwanted_cols = {
            'Дебет_начало', 'Кредит_начало', 'Дебет_конец', 
            'Кредит_конец', 'Исх.файл'
        }
        df = df.drop(columns=unwanted_cols, errors='ignore')
        df = df.set_index('Счет')
        
        # 4. Проверка наличия счетов в справочнике
        all_loads_df = DataLoader.load_reference_data('Выгрузки')
        
        osv_codes_set = set(df.index.str[:2].unique())
        ref_codes_set = set(all_loads_df['счет'].str[:2].unique())
        missing_codes = list(osv_codes_set - ref_codes_set)
        
        if missing_codes:
            # Формируем problem_data для ReferenceMismatchError
            problem_rows_mask = df.index.str[:2].isin(missing_codes)
            problem_data = (
                df.loc[problem_rows_mask]
                .reset_index()
                .rename(columns={'Счет': 'счет_осв'})
            )
            problem_data['код_отсутствующий_в_справочнике'] = problem_data['счет_осв'].str[:2]
            
            raise ReferenceMismatchError(
                message=(
                    f"Счета из общей ОСВ (файл {name_file}), "
                    f"которых нет в Справочники.xlsx (лист 'Выгрузки')"
                ),
                problem_data=problem_data,
                reference_name="Выгрузки",
                file_name=name_file,
                missing_codes=sorted(missing_codes)
            )
        
        logger.debug(f"Все счета из общей ОСВ (файл {name_file}) присутствуют в справочнике.")
        
        # 5. Сопоставление счетов (логика префиксов) - ТОЛЬКО ДЛЯ БАЛАНСА
        mapping_df = DataLoader.load_reference_data('Меппинг', **REFERENCE_CONFIGS['Меппинг'])
        # mapping_df = mapping_df[mapping_df['отчетность'] == "Баланс"]
        
        osv_codes = df.index.unique()
        mapping_codes_set = set(mapping_df['счет'].unique())
        
        # Словарь префиксов справочника
        prefix_dict = {}
        for code in mapping_codes_set:
            parts = code.split('.')
            for i in range(1, len(parts) + 1):
                prefix_dict.setdefault('.'.join(parts[:i]), []).append(code)
        
        # Анализ совпадений
        stats = {'total': len(osv_codes), 'full': 0, 'partial': 0, 'missing': 0}
        unmatched_details = []
        partial_matches = []
        
        for osv_code in osv_codes:
            if osv_code in mapping_codes_set:
                stats['full'] += 1
                continue
            
            # Поиск по префиксам от длинного к короткому
            matched_codes = []
            osv_parts = osv_code.split('.')
            for depth in range(len(osv_parts), 0, -1):
                prefix = '.'.join(osv_parts[:depth])
                if prefix in prefix_dict:
                    matched_codes = prefix_dict[prefix]
                    break
            
            if matched_codes:
                # Проверка: должно быть ровно одно совпадение
                if len(matched_codes) > 1:
                    raise ValueError(
                        f"Обнаружены несколько счетов в столбце 'Совпадающие счета из справочника': "
                        f"Счет из ОСВ='{osv_code}', "
                        f"Совпадающие счета={sorted(matched_codes)}. "
                        f"Проверьте справочник Меппинг — для каждого счета из ОСВ "
                        f"должно быть не более одного соответствия."
                    )
                
                stats['partial'] += 1
                partial_matches.append({
                    'Счет из ОСВ': osv_code,
                    'Совпадающие счета из справочника': matched_codes[0]
                })
            else:
                stats['missing'] += 1
                unmatched_details.append(osv_code)
        
        logger.debug(
            f"Статистика меппинга счетов - без совпадений: {stats['missing']} "
            f"{unmatched_details if unmatched_details else ''}"
        )
        
        # Создаем DataFrame с частичными совпадениями
        partially_matching_accounts_df = pd.DataFrame(partial_matches)
        if not partially_matching_accounts_df.empty:
            logger.debug(
                f"Найдено {len(partially_matching_accounts_df)} счетов с частичным совпадением"
            )
            # =========================================================================
        # =========================================================================
        # 6. ФОРМИРОВАНИЕ СПИСКА ВЫГРУЗОК ДЛЯ БАЛАНСА (ОСВ)
        # =========================================================================
        
        # Определяем валидные счета для баланса через process_account
        valid_accounts = df.index.map(process_account).dropna().unique()
        
        # ★ ИСПРАВЛЕНИЕ: фильтруем не только по счету, но и по типу регистра 'осв'
        # Это защищает от попадания отчета по проводкам (для ОПУ) в список ОСВ
        filtered_loads_balance = all_loads_df[
            all_loads_df['счет'].isin(valid_accounts) & 
            (all_loads_df['регистр'].str.lower() == 'осв')
        ].copy()
        
        filtered_loads_balance['Сокращенное Наименование компании'] = name_company
        filtered_loads_balance['Период Отчетности'] = period
        filtered_loads_balance['Тип регистра'] = 'осв'  # Маркер для Excel
        
        logger.debug(
            f"Для баланса отобрано {len(filtered_loads_balance)} ОСВ "
            f"по {len(valid_accounts)} счетам"
        )
        
        # =========================================================================
        # 7. ФОРМИРОВАНИЕ СПИСКА ОТЧЕТОВ ПО ПРОВОДКАМ ДЛЯ ОПУ
        # =========================================================================
        opu_prefixes = self._get_opu_accounts_to_export(df)
        
        if opu_prefixes:
            logger.info(
                f"Обнаружены обороты по счетам ОПУ: {sorted(opu_prefixes)}. "
                f"Формируем список карточек для выгрузки."
            )
            
            # ★ ИСПРАВЛЕНИЕ: фильтруем не только по префиксу, но и по типу регистра 'карточка'
            # Это защищает от попадания ОСВ в список карточек
            mask_opu_account = all_loads_df['счет'].str[:2].isin(opu_prefixes)
            mask_opu_register = all_loads_df['регистр'].str.lower() == 'отчет по проводкам'
            
            filtered_loads_opu = all_loads_df[mask_opu_account & mask_opu_register].copy()
            
            if filtered_loads_opu.empty:
                logger.warning(
                    f"⚠️ В справочнике 'Выгрузки' отсутствуют строки с "
                    f"регистром='отчет по проводкам' для счетов {sorted(opu_prefixes)}. "
                    f"Добавьте их в справочник для формирования списка карточек."
                )
                context.data['expected_card_filenames'] = []
                context.data['has_opu'] = False
                
                filtered_loads = filtered_loads_balance
            else:
                filtered_loads_opu['Сокращенное Наименование компании'] = name_company
                filtered_loads_opu['Период Отчетности'] = period
                filtered_loads_opu['Тип регистра'] = 'отчет по проводкам'  # Маркер для Excel
                
                # Формируем имена файлов для карточек
                card_filenames = self._format_card_filename_vectorized(filtered_loads_opu, format_file = 'txt')
                filtered_loads_opu['Имя файла для сохранения'] = card_filenames
                
                context.data['expected_card_filenames'] = card_filenames
                context.data['has_opu'] = True
                
                # Объединяем ОСВ и карточки в один DataFrame
                filtered_loads = pd.concat(
                    [filtered_loads_balance, filtered_loads_opu],
                    ignore_index=True
                )
                
                logger.info(
                    f"Для ОПУ необходимо выгрузить {len(card_filenames)} отчеты по проводкам: "
                    f"{sorted(filtered_loads_opu['счет'].unique())}"
                )
        else:
            # Нет счетов для ОПУ
            filtered_loads = filtered_loads_balance
            context.data['expected_card_filenames'] = []
            context.data['has_opu'] = False
            logger.info("Обороты по счетам ОПУ не обнаружены, сборка ОПУ будет пропущена")
        
        # =========================================================================
        # 8. ФОРМИРОВАНИЕ ИМЕН ФАЙЛОВ ДЛЯ ОСВ
        # =========================================================================
        mask_osv = filtered_loads['Тип регистра'] == 'осв'
        balance_filenames = format_filename_vectorized(filtered_loads[mask_osv])
        
        # Добавляем имя файла в DataFrame для удобства бухгалтера (только для ОСВ)
        filtered_loads.loc[mask_osv, 'Имя файла для сохранения'] = balance_filenames
        
        context.data['expected_filenames'] = balance_filenames
        context.data['mapping'] = mapping_df
        context.data['partially_matching_accounts_df'] = partially_matching_accounts_df
        
        # =========================================================================
        # 9. ФОРМИРОВАНИЕ СПИСКА СПЕЦОТЧЕТОВ (НЕОБЯЗАТЕЛЬНЫЕ)
        # =========================================================================
        special_reports_df = self._build_special_reports_list(name_company, period)
        
        # =========================================================================
        # 10. СОХРАНЕНИЕ В EXCEL (один лист для ОСВ + карточек + спецотчёты)
        # =========================================================================
        output_filename = f'Выгрузить_{name_company}_{period}.xlsx'
        output_path = OUTPUT_DATA_DIR / output_filename
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Лист 1: Обязательные выгрузки (ОСВ + карточки)
                filtered_loads.to_excel(
                    writer,
                    sheet_name='Обязательные выгрузки',
                    index=False
                )
                self._format_worksheet(writer.sheets['Обязательные выгрузки'])
                
                # Лист 2: Спецотчёты (необязательные)
                special_reports_df.to_excel(
                    writer,
                    sheet_name='Спецотчеты',
                    index=False
                )
                self._format_special_reports_sheet(
                    writer.sheets['Спецотчеты']
                )
                
        except PermissionError:
            error_msg = (
                f"❌ Не удалось сохранить файл '{output_path.name}': "
                f"файл открыт в Excel или другой программе.\n"
                f"📂 Путь: {output_path}\n"
                f"🔧 Закройте файл '{output_path.name}' и перезапустите программу."
            )
            logger.error(error_msg)
            raise PermissionError(error_msg) from None
        
        context.data['special_reports_filenames'] = special_reports_df['Имя файла для сохранения'].tolist()
        
        logger.info(
            f"Необходимо выгрузить {len(balance_filenames)} ОСВ и "
            f"{len(context.data.get('expected_card_filenames', []))} отчетов по проводкам."
        )
        logger.info(
            f"Подробности см. в {output_path.parent.name}/{output_path.name} "
            f"(2 листа: обязательные + спецотчеты)"
        )
        
        return context
    
    def _get_opu_accounts_to_export(self, df: pd.DataFrame) -> list:
        """
        Определяет префиксы счетов для выгрузки отчетов по проводкам под ОПУ.
        
        Критерий: счёт начинается с префиксов ОПУ (90, 91, 26, 44, 99)
        И есть ненулевой оборот (дебетовый ИЛИ кредитовый).
        
        Args:
            df: DataFrame общей ОСВ (с индексом 'Счет' — для совместимости)
                
        Returns:
            Список префиксов счетов (например, ['26', '90', '91'])
        """
        # ★ Используем исходный df из контекста (до set_index и удаления столбцов)
        original_df = getattr(self, '_original_osv_df', None)
        
        if original_df is None:
            logger.warning(
                "Исходный DataFrame с оборотами недоступен, "
                "невозможно определить счета для ОПУ"
            )
            return []
        

        accounts = original_df['Счет'].astype(str)
        
        # Фильтр по префиксам счетов ОПУ
        mask_opu_accounts = accounts.str.startswith(OPU_ACCOUNTS_PREFIXES, na=False)
        
        if not mask_opu_accounts.any():
            return []
        
        # Фильтр по наличию оборота (НЕ сальдо!)
        # Проверяем, что столбцы оборотов существуют
        has_debit = 'Дебет_оборот' in original_df.columns
        has_credit = 'Кредит_оборот' in original_df.columns
        
        if not has_debit and not has_credit:
            logger.warning(
                "В общей ОСВ отсутствуют столбцы 'Дебет_оборот' и 'Кредит_оборот'. "
                "Невозможно определить счета для ОПУ по оборотам."
            )
            return []
        
        # Формируем маску наличия оборота
        debit_turnover = original_df['Дебет_оборот'] if has_debit else 0
        credit_turnover = original_df['Кредит_оборот'] if has_credit else 0
        
        mask_has_turnover = (
            (pd.to_numeric(debit_turnover, errors='coerce').abs() > 0) | 
            (pd.to_numeric(credit_turnover, errors='coerce').abs() > 0)
        )
        
        # Счета для выгрузки отчетов по проводкам
        opu_accounts = accounts[mask_opu_accounts & mask_has_turnover]
        
        # Извлекаем префиксы (первые 2 символа)
        opu_prefixes = sorted(set(acc[:2] for acc in opu_accounts))
        
        return opu_prefixes
    
    
    def _get_original_osv_with_turnovers(self) -> pd.DataFrame:
        """
        Возвращает исходный DataFrame ОСВ с оборотами.
        
        К этому моменту в методе _process df уже очищен от оборотов,
        поэтому нужно получить оригинал.
        
        Хак: сохраняем ссылку на оригинал в начале _process.
        """
        return getattr(self, '_original_osv_df', None)
    
    
    @staticmethod
    def _format_card_filename_vectorized(df: pd.DataFrame, format_file = 'xlsx') -> list:
        """
        Формирует имена файлов для отчетов по проводкам.
        
        Формат: {company}_отчпровод_{account}_{period}_.xlsx
        
        Args:
            df: DataFrame с колонками:
                - 'Сокращенное Наименование компании'
                - 'счет'
                - 'Период Отчетности'
            format_file: str:
                xlsx - по умолчанию
                txt - для отчетов по проводкам (ОПУ) т.к. кол-во строк больше, чем в excel
        Returns:
            Список имён файлов
        """
        if df.empty:
            return []
        
        filenames = (
            df['Сокращенное Наименование компании'].astype(str) + '_' +
            'отчпровод' + '_' +
            df['счет'].astype(str) + '_' +
            df['Период Отчетности'].astype(str) + F'_.{format_file}'
        )
        
        return filenames.tolist()